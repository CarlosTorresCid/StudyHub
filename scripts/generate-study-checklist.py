#!/usr/bin/env python3
"""Genera exports/studyhub-checklist-estudio.xlsx a partir de src/data/public/<asignatura>/.

Lee unicamente los JSON publicos (configuracion.json, temas/*.json,
preguntas/banco-preguntas.json) y no modifica ningun archivo del proyecto.
Re-ejecutar este script regenera el Excel desde cero.
"""

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "src" / "data" / "public"
OUT_DIR = BASE_DIR / "exports"
OUT_FILE = OUT_DIR / "studyhub-checklist-estudio.xlsx"

SUBJECT_ORDER = ["smpc", "casi", "te", "iaic", "aamd", "isa"]
SHEET_NAMES = {
    "smpc": "SMPC",
    "casi": "CASI",
    "te": "TE",
    "iaic": "IAIC",
    "aamd": "AAMD",
    "isa": "ISA",
}

CHECK_EMPTY = "☐"  # ☐
CHECK_DONE = "☑"  # ☑

PRIORIDAD_SCORE = {"alta": 3, "media": 2, "baja": 1}
PRIORIDAD_OPCIONES = ["Alta", "Media", "Baja", "Sin datos"]

ANSWER_FIELDS = ["respuestaCorrecta", "respuestaModelo", "solucionOrientativa", "criteriosCorreccion"]

TEMARIO_HEADERS = [
    "Asignatura", "Tema ID", "Tema", "Título / nombre del tema", "Tiene contenido",
    "Nº conceptos / apartados", "Nº flashcards", "Nº preguntas asociadas", "Prioridad",
    "Estudiado", "Repasado", "Dominado", "Notas",
]
TEMARIO_CHECK_COLS = ["Estudiado", "Repasado", "Dominado"]

EXAMEN_HEADERS = [
    "Asignatura", "Pregunta ID", "Parte examen", "Tipo", "Tema principal",
    "Temas relacionados", "Grupo temático", "Origen", "Prioridad examen", "Modelo examen",
    "Apariciones / fuentes", "Enunciado resumido", "Tiene respuesta", "Requiere revisión",
    "Verificada", "Practicada", "Fallada", "Dominada", "Notas",
]
EXAMEN_CHECK_COLS = ["Practicada", "Fallada", "Dominada"]

RESUMEN_HEADERS = [
    "Asignatura", "Nº temas", "Nº temas con contenido", "Nº preguntas totales",
    "Nº preguntas tipo test", "Nº preguntas cortas", "Nº preguntas desarrollo",
    "Nº problemas/prácticas", "Nº preguntas con respuesta", "Nº preguntas sin respuesta",
    "Nº requieren revisión", "Prioridad general", "Notas",
]
RESUMEN_PARTES_HEADERS = ["Asignatura", "Parte examen", "Nº preguntas"]

COLUMN_WIDTHS = {
    "Asignatura": 12, "Tema ID": 10, "Tema": 8, "Título / nombre del tema": 50,
    "Tiene contenido": 14, "Nº conceptos / apartados": 12, "Nº flashcards": 12,
    "Nº preguntas asociadas": 12, "Prioridad": 10, "Estudiado": 11, "Repasado": 11,
    "Dominado": 11, "Notas": 35, "Pregunta ID": 26, "Parte examen": 22, "Tipo": 12,
    "Tema principal": 12, "Temas relacionados": 18, "Grupo temático": 28, "Origen": 14,
    "Prioridad examen": 14, "Modelo examen": 22, "Apariciones / fuentes": 38,
    "Enunciado resumido": 60, "Tiene respuesta": 13, "Requiere revisión": 14,
    "Verificada": 11, "Practicada": 11, "Fallada": 10, "Dominada": 11,
    "Nº temas": 9, "Nº temas con contenido": 12, "Nº preguntas totales": 12,
    "Nº preguntas tipo test": 12, "Nº preguntas desarrollo": 12,
    "Nº problemas/prácticas": 12, "Nº preguntas con respuesta": 12,
    "Nº preguntas sin respuesta": 12, "Nº requieren revisión": 12, "Prioridad general": 12,
}

WRAP_COLUMNS = {
    "Título / nombre del tema", "Notas", "Enunciado resumido", "Temas relacionados",
    "Grupo temático", "Apariciones / fuentes",
}

# ── Estilos ──────────────────────────────────────────────────────────────

THIN = Side(style="thin", color="C9C9C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FONT = Font(bold=True, color="1F1F1F", size=10)
ALT_FILL = PatternFill("solid", fgColor="F3F3F9")
EXAMEN_TITLE_FILL = PatternFill("solid", fgColor="595959")
EXAMEN_HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")


def lighten(hex_color, factor):
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    r, g, b = (round(c + (255 - c) * factor) for c in (r, g, b))
    return f"{r:02X}{g:02X}{b:02X}"


# ── Carga de datos ───────────────────────────────────────────────────────

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def extract_numero(tema_id):
    match = re.match(r"^tema-(\d+)$", tema_id or "", re.IGNORECASE)
    return int(match.group(1)) if match else None


def sort_temas(tema_ids):
    def key(tema_id):
        numero = extract_numero(tema_id)
        if numero is not None:
            return (0, numero, "")
        return (1, 0, tema_id)
    return sorted(tema_ids, key=key)


def has_value(v):
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    if isinstance(v, (list, dict)):
        return len(v) > 0
    return True  # numeros (incl. 0) y booleanos


def truncate(text, limit=120):
    text = re.sub(r"\s+", " ", (text or "").strip())
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def build_parte_maps(config):
    parte_nombre = {}
    tipo_to_parte = {}
    orden = []
    for bloque in config.get("estructuraExamen", []) or []:
        nombre = bloque.get("nombre", bloque.get("id", ""))
        parte_nombre[bloque.get("id")] = nombre
        orden.append(nombre)
        for tipo in bloque.get("tipos", []) or []:
            tipo_to_parte.setdefault(tipo, nombre)
    return parte_nombre, tipo_to_parte, orden


def parte_examen_nombre(q, parte_nombre, tipo_to_parte):
    parte_id = q.get("parteExamenId") or q.get("bloqueExamenId")
    if parte_id in parte_nombre:
        return parte_nombre[parte_id]
    tipo = q.get("tipo")
    if tipo in tipo_to_parte:
        return tipo_to_parte[tipo]
    return parte_id or "Sin clasificar"


def load_subject(sid):
    subject_dir = DATA_DIR / sid
    config = {}
    config_path = subject_dir / "configuracion.json"
    if config_path.exists():
        config = load_json(config_path)

    contents = {}
    temas_dir = subject_dir / "temas"
    if temas_dir.exists():
        for f in sorted(temas_dir.glob("*.json")):
            data = load_json(f)
            tema_id = data.get("temaId")
            if tema_id:
                contents[tema_id] = data

    questions = []
    bank_path = subject_dir / "preguntas" / "banco-preguntas.json"
    if bank_path.exists():
        data = load_json(bank_path)
        if isinstance(data, list):
            questions = data
        elif data:
            questions = [data]

    return config, contents, questions


# ── Calculo de filas ─────────────────────────────────────────────────────

def tema_priority_score(asociadas):
    # Combina nº de preguntas, prioridadExamen, apariciones/numeroApariciones,
    # presencia en examenes reales y nº de modelos distintos en los que aparece el tema.
    score = len(asociadas)
    modelos = set()
    hay_examen_real = False
    for q in asociadas:
        score += PRIORIDAD_SCORE.get((q.get("prioridadExamen") or "").lower(), 0) * 2
        if q.get("origen") == "examen_real":
            hay_examen_real = True
        apariciones = q.get("apariciones")
        if isinstance(apariciones, list):
            score += len(apariciones)
        num_ap = q.get("numeroApariciones")
        if isinstance(num_ap, (int, float)):
            score += num_ap
        for key in ("modeloExamenId", "carpetaPrincipalId"):
            if q.get(key):
                modelos.add(q[key])
        for m in q.get("modelos") or []:
            modelos.add(m)
    score += len(modelos)
    if hay_examen_real:
        score += 2
    return score


def assign_priorities(scores_by_tema):
    # Ranking relativo por terciles dentro de cada asignatura: con thresholds
    # absolutos casi todos los temas con preguntas caían en "Alta" y la columna
    # no servia para priorizar. Los temas sin preguntas quedan "Sin datos".
    con_preguntas = [(t, s) for t, s in scores_by_tema.items() if s > 0]
    n = len(con_preguntas)
    ordenados = sorted(con_preguntas, key=lambda x: x[1], reverse=True)
    prioridades = {}
    for idx, (tema_id, _) in enumerate(ordenados):
        frac = idx / n
        if frac < 1 / 3:
            prioridades[tema_id] = "Alta"
        elif frac < 2 / 3:
            prioridades[tema_id] = "Media"
        else:
            prioridades[tema_id] = "Baja"
    for tema_id, score in scores_by_tema.items():
        if score == 0:
            prioridades[tema_id] = "Sin datos"
    return prioridades


def build_temario_rows(sid, abrev, contents, questions):
    tema_ids = set(contents.keys())
    for q in questions:
        if q.get("temaPrincipalId"):
            tema_ids.add(q["temaPrincipalId"])
        for t in q.get("temaIds") or []:
            tema_ids.add(t)

    tema_ids = sort_temas(tema_ids)
    asociadas_por_tema = {}
    for tema_id in tema_ids:
        asociadas_por_tema[tema_id] = [
            q for q in questions
            if q.get("temaPrincipalId") == tema_id or tema_id in (q.get("temaIds") or [])
        ]

    scores = {t: tema_priority_score(asociadas) for t, asociadas in asociadas_por_tema.items()}
    prioridades = assign_priorities(scores)

    rows = []
    for tema_id in tema_ids:
        content = contents.get(tema_id)
        titulo = content.get("titulo") if content else tema_id
        tiene_contenido = "Sí" if content else "No"

        n_secciones = len((content or {}).get("apuntes", {}).get("secciones", []) or [])
        n_conceptos_clave = len((content or {}).get("conceptosClave", []) or [])
        n_flashcards = len((content or {}).get("flashcards", []) or [])

        rows.append({
            "Asignatura": abrev,
            "Tema ID": tema_id,
            "Tema": extract_numero(tema_id) if extract_numero(tema_id) is not None else "",
            "Título / nombre del tema": titulo,
            "Tiene contenido": tiene_contenido,
            "Nº conceptos / apartados": n_secciones + n_conceptos_clave,
            "Nº flashcards": n_flashcards,
            "Nº preguntas asociadas": len(asociadas_por_tema[tema_id]),
            "Prioridad": prioridades[tema_id],
            "Estudiado": CHECK_EMPTY,
            "Repasado": CHECK_EMPTY,
            "Dominado": CHECK_EMPTY,
            "Notas": "",
        })

    return rows


def build_examen_rows(sid, abrev, questions, parte_nombre, tipo_to_parte):
    rows = []
    for i, q in enumerate(questions):
        tema_principal = q.get("temaPrincipalId", "")
        temas_relacionados = [t for t in (q.get("temaIds") or []) if t != tema_principal]

        modelo_examen = (
            q.get("modeloExamenNombre")
            or q.get("modeloExamenId")
            or q.get("carpetaPrincipalNombre")
            or (", ".join(q.get("modelos") or []) if q.get("modelos") else "")
            or ""
        )

        fuente_parts = []
        for key in ("fuentes", "apariciones", "etiquetasModelo"):
            v = q.get(key)
            if isinstance(v, list):
                fuente_parts.extend(str(x) for x in v)
            elif has_value(v):
                fuente_parts.append(str(v))
        for key in ("modeloExamenNombre", "modeloExamenId"):
            v = q.get(key)
            if has_value(v):
                fuente_parts.append(str(v))
        num_ap = q.get("numeroApariciones")
        if isinstance(num_ap, (int, float)) and num_ap:
            fuente_parts.append(f"{num_ap} apariciones")
        apariciones_fuentes = "; ".join(dict.fromkeys(fuente_parts))

        prioridad_examen = q.get("prioridadExamen")
        prioridad_examen = prioridad_examen.capitalize() if prioridad_examen else "Sin datos"

        tiene_respuesta = "Sí" if any(has_value(q.get(k)) for k in ANSWER_FIELDS) else "No"

        rows.append({
            "Asignatura": abrev,
            "Pregunta ID": q.get("id") or f"{sid}-{i + 1:03d}",
            "Parte examen": parte_examen_nombre(q, parte_nombre, tipo_to_parte),
            "Tipo": q.get("tipo", ""),
            "Tema principal": tema_principal,
            "Temas relacionados": ", ".join(temas_relacionados),
            "Grupo temático": q.get("grupoTematico", ""),
            "Origen": q.get("origen", ""),
            "Prioridad examen": prioridad_examen,
            "Modelo examen": modelo_examen,
            "Apariciones / fuentes": apariciones_fuentes,
            "Enunciado resumido": truncate(q.get("enunciado", "")),
            "Tiene respuesta": tiene_respuesta,
            "Requiere revisión": "Sí" if q.get("requiereRevision") else "No",
            "Verificada": "Sí" if q.get("verificada") else "No",
            "Practicada": CHECK_EMPTY,
            "Fallada": CHECK_EMPTY,
            "Dominada": CHECK_EMPTY,
            "Notas": "",
        })

    return rows


TEST_TIPOS = {"test", "verdadero_falso"}
PRACTICA_TIPOS = {"practica", "problema"}


def build_resumen_row(abrev, temario_rows, examen_rows):
    tipo_counts = Counter(r["Tipo"] for r in examen_rows)
    n_test = sum(c for t, c in tipo_counts.items() if t in TEST_TIPOS)
    n_corta = tipo_counts.get("corta", 0)
    n_desarrollo = tipo_counts.get("desarrollo", 0)
    n_practica = sum(c for t, c in tipo_counts.items() if t in PRACTICA_TIPOS)
    n_con_respuesta = sum(1 for r in examen_rows if r["Tiene respuesta"] == "Sí")
    n_revision = sum(1 for r in examen_rows if r["Requiere revisión"] == "Sí")

    return {
        "Asignatura": abrev,
        "Nº temas": len(temario_rows),
        "Nº temas con contenido": sum(1 for r in temario_rows if r["Tiene contenido"] == "Sí"),
        "Nº preguntas totales": len(examen_rows),
        "Nº preguntas tipo test": n_test,
        "Nº preguntas cortas": n_corta,
        "Nº preguntas desarrollo": n_desarrollo,
        "Nº problemas/prácticas": n_practica,
        "Nº preguntas con respuesta": n_con_respuesta,
        "Nº preguntas sin respuesta": len(examen_rows) - n_con_respuesta,
        "Nº requieren revisión": n_revision,
        "Prioridad general": "",
        "Notas": "",
    }


def build_partes_rows(abrev, examen_rows, orden_partes):
    counts = Counter(r["Parte examen"] for r in examen_rows)
    rows = []
    for nombre in orden_partes:
        if nombre in counts:
            rows.append({"Asignatura": abrev, "Parte examen": nombre, "Nº preguntas": counts.pop(nombre)})
    for nombre, n in counts.items():
        rows.append({"Asignatura": abrev, "Parte examen": nombre, "Nº preguntas": n})
    return rows


# ── Escritura de Excel ───────────────────────────────────────────────────

def style_title_cell(cell, fill, font=TITLE_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header_cell(cell, fill):
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def write_data_row(ws, row_idx, headers, data, alt, wrap_cols):
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=c, value=data.get(header, ""))
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=header in wrap_cols,
        )
        if alt:
            cell.fill = ALT_FILL


def add_checklist_validation(ws, col_idx, first_row, last_row):
    if last_row < first_row:
        return
    dv = DataValidation(type="list", formula1=f'"{CHECK_EMPTY},{CHECK_DONE}"', allow_blank=True)
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col_idx)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def set_column_widths(ws, max_cols, *header_lists):
    for c in range(1, max_cols + 1):
        widths = []
        for headers in header_lists:
            if c <= len(headers):
                widths.append(COLUMN_WIDTHS.get(headers[c - 1], 12))
        ws.column_dimensions[get_column_letter(c)].width = max(widths) if widths else 12


def write_subject_sheet(wb, sid, nombre, abrev, color, temario_rows, examen_rows):
    ws = wb.create_sheet(title=SHEET_NAMES[sid])

    title_fill = PatternFill("solid", fgColor=color.lstrip("#"))
    header_fill = PatternFill("solid", fgColor=lighten(color, 0.75))

    n_temario = len(TEMARIO_HEADERS)
    n_examen = len(EXAMEN_HEADERS)
    max_cols = max(n_temario, n_examen)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_temario)
    style_title_cell(ws.cell(row=row, column=1, value=f"A) TEMARIO — {nombre}"), title_fill)
    row += 1

    for c, header in enumerate(TEMARIO_HEADERS, start=1):
        style_header_cell(ws.cell(row=row, column=c, value=header), header_fill)
    row += 1

    temario_first = row
    for i, data in enumerate(temario_rows):
        write_data_row(ws, row, TEMARIO_HEADERS, data, i % 2 == 1, WRAP_COLUMNS)
        row += 1
    temario_last = row - 1

    row += 1  # fila separadora

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_examen)
    style_title_cell(ws.cell(row=row, column=1, value=f"B) EXAMEN — {nombre}"), EXAMEN_TITLE_FILL)
    row += 1

    examen_header_row = row
    for c, header in enumerate(EXAMEN_HEADERS, start=1):
        style_header_cell(ws.cell(row=row, column=c, value=header), EXAMEN_HEADER_FILL)
    row += 1

    examen_first = row
    for i, data in enumerate(examen_rows):
        write_data_row(ws, row, EXAMEN_HEADERS, data, i % 2 == 1, WRAP_COLUMNS)
        row += 1
    examen_last = row - 1

    set_column_widths(ws, max_cols, TEMARIO_HEADERS, EXAMEN_HEADERS)

    for col_name in TEMARIO_CHECK_COLS:
        add_checklist_validation(ws, TEMARIO_HEADERS.index(col_name) + 1, temario_first, temario_last)
    for col_name in EXAMEN_CHECK_COLS:
        add_checklist_validation(ws, EXAMEN_HEADERS.index(col_name) + 1, examen_first, examen_last)

    if examen_last >= examen_first:
        ws.auto_filter.ref = f"A{examen_header_row}:{get_column_letter(n_examen)}{examen_last}"

    ws.freeze_panes = f"B{examen_header_row + 1}"

    return {
        "n_temario": len(temario_rows),
        "n_examen": len(examen_rows),
    }


def write_resumen_sheet(wb, resumen_rows, partes_rows):
    ws = wb.create_sheet(title="Resumen", index=0)

    title_fill = PatternFill("solid", fgColor="44546A")
    header_fill = PatternFill("solid", fgColor="D6DCE5")

    n_cols = len(RESUMEN_HEADERS)
    row = 1
    for c, header in enumerate(RESUMEN_HEADERS, start=1):
        style_header_cell(ws.cell(row=row, column=c, value=header), header_fill)
    row += 1

    first = row
    for i, data in enumerate(resumen_rows):
        write_data_row(ws, row, RESUMEN_HEADERS, data, i % 2 == 1, WRAP_COLUMNS)
        row += 1
    last = row - 1

    if last >= first:
        col_idx = RESUMEN_HEADERS.index("Prioridad general") + 1
        dv = DataValidation(type="list", formula1=f'"{",".join(PRIORIDAD_OPCIONES)}"', allow_blank=True)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}{first}:{col_letter}{last}")

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(RESUMEN_PARTES_HEADERS))
    style_title_cell(ws.cell(row=row, column=1, value="Preguntas por parte de examen"), title_fill)
    row += 1

    for c, header in enumerate(RESUMEN_PARTES_HEADERS, start=1):
        style_header_cell(ws.cell(row=row, column=c, value=header), header_fill)
    row += 1

    for i, data in enumerate(partes_rows):
        write_data_row(ws, row, RESUMEN_PARTES_HEADERS, data, i % 2 == 1, WRAP_COLUMNS)
        row += 1

    set_column_widths(ws, max(n_cols, len(RESUMEN_PARTES_HEADERS)), RESUMEN_HEADERS, RESUMEN_PARTES_HEADERS)

    if last >= first:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{last}"
    ws.freeze_panes = "A2"


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    asignaturas = {a["id"]: a for a in load_json(DATA_DIR / "asignaturas.json")}

    OUT_DIR.mkdir(exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # se sustituye por la hoja "Resumen" al final

    resumen_rows = []
    partes_rows = []
    warnings = []

    for sid in SUBJECT_ORDER:
        subject_dir = DATA_DIR / sid
        info = asignaturas.get(sid, {})
        nombre = info.get("nombre", sid.upper())
        abrev = info.get("abreviatura", sid.upper())
        color = info.get("color", "#4F7CAC")

        if not subject_dir.exists():
            warnings.append(f"{sid}: no existe la carpeta {subject_dir}")
            config, contents, questions = {}, {}, []
        else:
            config, contents, questions = load_subject(sid)
            if not contents:
                warnings.append(f"{sid}: no se encontraron temas/*.json")
            if not questions:
                warnings.append(f"{sid}: no se encontró banco-preguntas.json o está vacío")

        parte_nombre, tipo_to_parte, orden_partes = build_parte_maps(config)

        temario_rows = build_temario_rows(sid, abrev, contents, questions)
        examen_rows = build_examen_rows(sid, abrev, questions, parte_nombre, tipo_to_parte)

        write_subject_sheet(wb, sid, nombre, abrev, color, temario_rows, examen_rows)

        resumen_rows.append(build_resumen_row(abrev, temario_rows, examen_rows))
        partes_rows.extend(build_partes_rows(abrev, examen_rows, orden_partes))

    write_resumen_sheet(wb, resumen_rows, partes_rows)

    wb.save(OUT_FILE)

    print(f"Excel generado: {OUT_FILE.relative_to(BASE_DIR)}")
    print("Hojas: " + ", ".join(["Resumen"] + [SHEET_NAMES[s] for s in SUBJECT_ORDER]))
    for r in resumen_rows:
        print(f"  {r['Asignatura']}: {r['Nº temas']} temas, {r['Nº preguntas totales']} preguntas")
    if warnings:
        print("Advertencias:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
